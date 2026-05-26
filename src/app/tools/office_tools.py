import json
from collections.abc import Iterable
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from langchain_core.tools import tool
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.service.generated_file_service import register_generated_file, resolve_generated_file_path


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
INVALID_SHEET_NAME_CHARS = set("[]:*?/\\")
MAX_SHEET_NAME_LENGTH = 31


def _ok(**data: Any) -> str:
    return json.dumps({"success": True, **data}, ensure_ascii=False, indent=2, default=str)


def _error(message: str) -> str:
    return json.dumps({"success": False, "error": message}, ensure_ascii=False, indent=2)


def _resolve_excel_path(file_path: str) -> Path:
    if not file_path or not file_path.strip():
        raise ValueError("file_path 不能为空")

    path = resolve_generated_file_path(file_path, default_suffix=".xlsx")

    if path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("仅支持 .xlsx 或 .xlsm 文件")

    return path


def _success_response(message: str, file_path: Path, sheet: Worksheet, **extra: Any) -> str:
    file_info = _register_excel_file(file_path)
    return _ok(
        message=message,
        file_path=str(file_path),
        sheet_name=sheet.title,
        file=file_info,
        files=[file_info],
        **extra,
    )


def _normalize_sheet_name(sheet_name: str, fallback: str = "Sheet1") -> str:
    name = (sheet_name or fallback).strip() or fallback
    for char in INVALID_SHEET_NAME_CHARS:
        name = name.replace(char, "_")
    return name[:MAX_SHEET_NAME_LENGTH]


def _get_or_create_sheet(workbook: OpenpyxlWorkbook, sheet_name: str = "") -> Worksheet:
    if sheet_name:
        normalized_name = _normalize_sheet_name(sheet_name)
        if normalized_name not in workbook.sheetnames:
            workbook.create_sheet(normalized_name)
        return workbook[normalized_name]
    return workbook.active


def _select_existing_sheet(workbook: OpenpyxlWorkbook, sheet_name: str = "") -> Worksheet:
    if not sheet_name:
        return workbook.active

    normalized_name = _normalize_sheet_name(sheet_name)
    if normalized_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在: {normalized_name}")
    return workbook[normalized_name]


def _get_sheet_for_new_workbook(workbook: OpenpyxlWorkbook, sheet_name: str = "") -> Worksheet:
    sheet = workbook.active
    if sheet_name:
        sheet.title = _normalize_sheet_name(sheet_name)
    return sheet


def _coerce_rows(rows: Iterable[Any] | None) -> list[list[Any]]:
    if rows is None:
        return []
    if not isinstance(rows, Iterable) or isinstance(rows, (str, bytes, dict)):
        raise ValueError("rows/data 必须是二维数组，例如 [[\"姓名\", \"分数\"], [\"小星\", 98]]")

    normalized_rows = []
    for row in rows:
        if isinstance(row, Iterable) and not isinstance(row, (str, bytes, dict)):
            normalized_rows.append(list(row))
        else:
            normalized_rows.append([row])
    return normalized_rows


def _clear_sheet(sheet: Worksheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)


def _write_rows(sheet: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _apply_header_style(sheet: Worksheet) -> None:
    if sheet.max_row < 1:
        return

    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True, color="1F2937")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def _format_sheet(sheet: Worksheet, freeze_header: bool = True, auto_width: bool = True) -> None:
    if freeze_header and sheet.max_row >= 1:
        sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
        _apply_header_style(sheet)
    if auto_width:
        _auto_fit_columns(sheet)


def _save_workbook(workbook: OpenpyxlWorkbook, file_path: Path) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)


def _register_excel_file(file_path: Path) -> dict[str, Any]:
    return register_generated_file(file_path, category="excel")


def _load_existing_workbook(file_path: Path) -> OpenpyxlWorkbook:
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    return load_workbook(file_path)


@contextmanager
def _editable_workbook(file_path: Path, *, create_if_missing: bool = True):
    if file_path.exists():
        workbook = load_workbook(file_path)
    elif create_if_missing:
        workbook = Workbook()
    else:
        raise FileNotFoundError(f"文件不存在: {file_path}")

    yield workbook
    _save_workbook(workbook, file_path)


def _append_table(
    sheet: Worksheet,
    headers: list[Any] | None,
    rows: list[list[Any]] | None,
) -> int:
    rows_written = 0
    if headers:
        sheet.append(list(headers))
        rows_written += 1

    data_rows = _coerce_rows(rows)
    _write_rows(sheet, data_rows)
    return rows_written + len(data_rows)


def _write_sheet_data(
    file_path: str,
    sheet_name: str,
    headers: list[Any] | None,
    rows: list[list[Any]] | None,
    replace: bool,
    auto_width: bool,
) -> str:
    path = _resolve_excel_path(file_path)
    with _editable_workbook(path) as workbook:
        sheet = (
            _get_or_create_sheet(workbook, sheet_name)
            if path.exists()
            else _get_sheet_for_new_workbook(workbook, sheet_name)
        )

        if replace:
            _clear_sheet(sheet)

        rows_written = _append_table(sheet, headers, rows)
        _format_sheet(sheet, auto_width=auto_width)

    logger.info("Excel 工作表已写入: {} -> {}", path, sheet.title)
    return _success_response(
        message="Excel 工作表已写入",
        rows_written=rows_written,
        mode="replace" if replace else "append",
        file_path=path,
        sheet=sheet,
    )


@tool(
    description=(
        "生成 Excel 工作簿。file_path: 输出路径（支持相对/绝对路径，默认补 .xlsx）；"
        "sheet_name: 工作表名称；headers: 表头数组；rows: 数据二维数组；"
        "overwrite: 文件存在时是否覆盖。返回保存路径和写入行数。"
    )
)
def excel_create_workbook(
    file_path: str,
    sheet_name: str = "Sheet1",
    headers: list[Any] | None = None,
    rows: list[list[Any]] | None = None,
    overwrite: bool = True,
    auto_width: bool = True,
) -> str:
    """生成新的 Excel 工作簿。"""
    try:
        path = _resolve_excel_path(file_path)
        if path.exists() and not overwrite:
            return _error(f"文件已存在，如需覆盖请设置 overwrite=True: {path}")

        workbook = Workbook()
        sheet = _get_sheet_for_new_workbook(workbook, sheet_name)

        rows_written = _append_table(sheet, headers, rows)
        _format_sheet(sheet, auto_width=auto_width)
        _save_workbook(workbook, path)

        logger.info("Excel 工作簿已生成: {}", path)
        return _success_response(
            message="Excel 工作簿已生成",
            rows_written=rows_written,
            file_path=path,
            sheet=sheet,
        )
    except Exception as e:
        logger.error("生成 Excel 工作簿失败: {}", e)
        return _error(str(e))


@tool(
    description=(
        "写入或替换 Excel 工作表数据。file_path: Excel 路径；sheet_name: 工作表名称，"
        "不存在会自动创建；headers: 表头数组；rows: 数据二维数组；"
        "replace: True 表示清空该工作表后重写，False 表示从末尾追加。"
    )
)
def excel_write_sheet(
    file_path: str,
    sheet_name: str = "Sheet1",
    headers: list[Any] | None = None,
    rows: list[list[Any]] | None = None,
    replace: bool = True,
    auto_width: bool = True,
) -> str:
    """写入、替换或追加 Excel 工作表数据。"""
    try:
        return _write_sheet_data(
            file_path=file_path,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            replace=replace,
            auto_width=auto_width,
        )
    except Exception as e:
        logger.error("写入 Excel 工作表失败: {}", e)
        return _error(str(e))


@tool(
    description=(
        "向 Excel 工作表追加多行数据。file_path: Excel 路径；sheet_name: 工作表名称；"
        "rows: 要追加的二维数组，例如 [[\"张三\", 90], [\"李四\", 85]]。"
    )
)
def excel_append_rows(
    file_path: str,
    sheet_name: str = "Sheet1",
    rows: list[list[Any]] | None = None,
    auto_width: bool = True,
) -> str:
    """向已有或新建的 Excel 工作表追加行。"""
    try:
        return _write_sheet_data(
            file_path=file_path,
            sheet_name=sheet_name,
            headers=None,
            rows=rows,
            replace=False,
            auto_width=auto_width,
        )
    except Exception as e:
        logger.error("追加 Excel 行失败: {}", e)
        return _error(str(e))


@tool(
    description=(
        "编辑 Excel 单元格。file_path: Excel 路径；sheet_name: 工作表名称；"
        "cells: 单元格和值的映射，例如 {\"A1\": \"姓名\", \"B2\": 98}；"
        "create_if_missing: 文件不存在时是否自动创建。"
    )
)
def excel_update_cells(
    file_path: str,
    sheet_name: str = "Sheet1",
    cells: dict[str, Any] | None = None,
    create_if_missing: bool = True,
    auto_width: bool = True,
) -> str:
    """批量更新 Excel 单元格。"""
    try:
        path = _resolve_excel_path(file_path)
        if not cells:
            return _error("cells 不能为空，例如 {\"A1\": \"姓名\", \"B2\": 98}")

        with _editable_workbook(path, create_if_missing=create_if_missing) as workbook:
            sheet = (
                _get_or_create_sheet(workbook, sheet_name)
                if path.exists()
                else _get_sheet_for_new_workbook(workbook, sheet_name)
            )
            for coordinate, value in cells.items():
                sheet[str(coordinate)] = value

            _format_sheet(sheet, auto_width=auto_width)

        logger.info("Excel 单元格已更新: {} -> {}", path, sheet.title)
        return _success_response(
            message="Excel 单元格已更新",
            cells_updated=len(cells),
            file_path=path,
            sheet=sheet,
        )
    except Exception as e:
        logger.error("编辑 Excel 单元格失败: {}", e)
        return _error(str(e))


@tool(
    description=(
        "读取 Excel 工作表预览。file_path: Excel 路径；sheet_name: 工作表名称，"
        "为空则读取活动工作表；max_rows/max_cols 控制返回范围。"
    )
)
def excel_read_sheet(
    file_path: str,
    sheet_name: str = "",
    max_rows: int = 20,
    max_cols: int = 20,
) -> str:
    """读取 Excel 工作表内容预览。"""
    try:
        path = _resolve_excel_path(file_path)
        workbook = _load_existing_workbook(path)
        sheet = _select_existing_sheet(workbook, sheet_name)

        actual_max_row = sheet.max_row
        actual_max_col = sheet.max_column
        row_limit = max(1, min(max_rows, 200, actual_max_row))
        col_limit = max(1, min(max_cols, 100, actual_max_col))
        rows = []
        for row in sheet.iter_rows(max_row=row_limit, max_col=col_limit, values_only=True):
            rows.append(list(row))

        return _ok(
            message="Excel 工作表读取成功",
            file_path=str(path),
            sheet_name=sheet.title,
            sheet_names=workbook.sheetnames,
            max_row=actual_max_row,
            max_column=actual_max_col,
            preview_rows=rows,
        )
    except Exception as e:
        logger.error("读取 Excel 工作表失败: {}", e)
        return _error(str(e))
